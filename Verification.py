import cv2
from pyzbar import pyzbar
url = "http://192.168.8.233:8080/video"
texxt=""
cap = cv2.VideoCapture(url)
while True:
    ret, frame = cap.read()
    frame = cv2.resize(frame, (0, 0),fx=0.50, fy=0.50)
    cv2.putText(frame, "Press q to close camera", (10,10),cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 100), 1)
    barcodes = pyzbar.decode(frame)
    for barcode in barcodes:
        (x, y, w, h) = barcode.rect
        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 1)
        barcodeData = barcode.data.decode("utf-8")
        barcodeType =barcode.type
        text = "{}".format(barcodeData)
        texxt=text
        cv2.putText(frame, text, (x, y - 10),cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 1)
    cv2.imshow("Barcode Scanner",frame)
    if cv2.waitKey(1)==ord('q'):
        break
    if texxt:
        break
print(texxt)
texxt=int(texxt)
cap.release()
cv2.destroyAllWindows()
import openpyxl
wb=openpyxl.load_workbook("F://entries.xlsx")
sh=wb['Sheet1']
row=sh.max_row
key=0
no=0
for i in range(3,row+1):
    if(sh.cell(i,9).value==int(texxt)):
        key=1
        sh.cell(row=i,column=7,value=sh.cell(i,7).value+1)
        if (sh.cell(i,7).value>1):
            print("ReEntry!!!!!!!!")
        wb.save("F://entries.xlsx")    #"F://entries.xlsx"
        break
if(key==0):
    print("Invalid Entry")
else:
    print("success...")
wb.close()
