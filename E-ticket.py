# import cv2
# from pyzbar import pyzbar
# url = "http://192.168.121.115:8080/video"
# texxt=""
# cap = cv2.VideoCapture(url)
# while True:
#     ret, frame = cap.read()
#     frame = cv2.resize(frame, (0, 0),fx=0.50, fy=0.50)
#     cv2.putText(frame, "Press q to close camera", (10,10),cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 100), 1)
#     barcodes = pyzbar.decode(frame)
#     for barcode in barcodes:
#         (x, y, w, h) = barcode.rect
#         cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 1)
#         barcodeData = barcode.data.decode("utf-8")
#         barcodeType =barcode.type
#         text = "{}".format(barcodeData)
#         texxt=text
#         cv2.putText(frame, text, (x, y - 10),cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 1)
#     cv2.imshow("Barcode Scanner",frame)
#     if cv2.waitKey(1)==ord('q'):
#         break
#     if texxt:
#         break
# print(texxt)
# cap.release()
# cv2.destroyAllWindows()
# import openpyxl
# wb=openpyxl.load_workbook("F:\list.xlsx")
# text=113400
# sh=wb["3rd Year Student List"]
# row=sh.max_row
# key=0
# for i in range(3,row+1):
#     if(sh.cell(i,9).value==text):
#         key=1
#         sh.cell(row=i,column=7,value=sh.cell(i,7).value+1)
#         wb.save("F:\list.xlsx")
#         break
# if(key==0):
#     print("Invalid Entry")
# else:
#     print("success...")

# import qrcode
# qr=111002
# for i in range(1,188):
#     img=qrcode.make(qr)
#     qr=qr+198
#     temp=str(i)
#     img.save("F://qrcodes//"+temp+".jpg")

# from PIL import Image
# import qrcode
# qr=111002
# for i in range(1,188):
#     img=qrcode.make(qr)
#     qr=qr+198
#     ticket=Image.open("C://Users//Dell//Downloads//ticket.png")
#     x,y=1640,60
#     area=(x,y,x+290,y+290)
#     ticket.paste(img,area)
#     ticketWithQr=ticket
#     temp=str(i)
#     ticket.save("F://qrcodes//"+temp+".png")
#

import pandas as pd
import smtplib
import fsspec
e=pd.read_excel("F://maildemo.xlsx")
emails=e['Email'].values
server=smtplib.SMTP("smtp.gmail.com",587)
server.starttls()
server.login("eticket.Ece.Farewell22.ABESEC@gmail.com","eticket.Ece.Farewell22")
msg="hello bro"
subject="Bhai"
body="Subject:{}\n\n{}".format(subject,msg)
for em in emails:
    server.sendmail("eticket.Ece.Farewell22.ABESEC@gmail.com",em,body)


# import smtplib
# import imghdr
# from email.message import EmailMessage
#
# Sender_Email = "eticket.Ece.Farewell22.ABESEC@gmail.com"
# Reciever_Email = "navneet.sharma@abes.ac.in"
# Password = "eticket.Ece.Farewell22"
#
# newMessage = EmailMessage()
# newMessage['Subject'] = "Sir Check This!"
# newMessage['From'] = Sender_Email
# newMessage['To'] = Reciever_Email
# newMessage.set_content('Let me know what you think. Image attached!')
#
# with open("F://qrcodes//41.png", 'rb') as f:
#     image_data = f.read()
#     image_type = imghdr.what(f.name)
#     image_name = f.name
#
# newMessage.add_attachment(image_data, maintype='image', subtype=image_type, filename=image_name)
#
# with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
#     smtp.login(Sender_Email, Password)
#     smtp.send_message(newMessage)





import pandas as pd
import fsspec
import openpyxl
import smtplib
import imghdr
from email.message import EmailMessage

def sendtic(rec,name,img):
    Sender_Email = "eticket.Ece.Farewell22.ABESEC@gmail.com"
    Reciever_Email = rec
    Password = "eticket.Ece.Farewell22"

    newMessage = EmailMessage()
    newMessage['Subject'] = "e-Tickets #YAADEIN'22 ECE-Farewell"
    newMessage['From'] = Sender_Email
    newMessage['To'] = Reciever_Email
    number=9026750548
    emailid="tushar.19b311179@abes.ac.in"
    newMessage.set_content("Dear {},\n\nGreetings of the day!!\nThere is no entry without the tickets, so you all are here by requested to download it beforehand.\n\n\nFor any query please contact: \nTushar Tegwani \n(ECE Branch 3rd Year) \n{} \n{}".format(name,number,emailid))

    with open("F://qrcodes//"+str(img)+".png", 'rb') as f:
        image_data = f.read()
        image_type = imghdr.what(f.name)
        image_name = "e-ticket_"+name+".png"

    newMessage.add_attachment(image_data, maintype='image', subtype=image_type, filename=image_name)

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(Sender_Email, Password)
        smtp.send_message(newMessage)

e=pd.read_excel("F://list.xlsx")
emails=e['EMAIL'].values
names=e['Name'].values
ticno=0
for em in emails:
    sendtic(em,names[ticno],ticno+1)
    ticno+=1
